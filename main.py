import openpyxl
from openpyxl import load_workbook
import datetime
import shutil
import string



def read_data(fileName):
	'''
	Reads in a text file, parses it into regular 2x2 array

	Args:
		fileName: a string representing the file where all data is copied into.
			Ex: 'headcount_data.txt'
	'''
	lines = []
	with open(fileName) as f:
		lines = f.readlines()

	# get sunday date info
	startDate = lines[0].split(": ")[1]
	lines = lines[2::]

	# get totals info
	totals = get_totals_info(lines)
	lines = lines[len(totals)+1::]

	(olderGroup, preKGroup, olderTotals, preKTotals) = get_groups(lines)

	verify_totals(totals, olderTotals, preKTotals)

	writeToSheets(totals, olderTotals, preKTotals, olderGroup, preKGroup, startDate)



def get_totals_info(lines):
	'''
	Helper function to get the program totals out of the text file,
	store it, and remove the line from the list of text lines

	Args:
		lines: an array representing the lines of the text file
	'''
	totals = []
	i = 0
	while "===" not in lines[i]:
		day = lines[i].strip()
		day = day.split(": ")
		totals.append(day)
		i += 1
		if i >= len(lines):
			raise Exception("No horizontal line break found (looks like `========`). Likely cause: text file is formatted incorrectly.")

	return totals



def verify_totals(totals, olderTotals, preKTotals):
	'''
	Checks that the totals add up properly

	Args:
		totals: a (string, int) array representing the number of kids present
			in the overall program each day
		olderTotals: an int array representing the total #older kids each day
		preKTotals: an int array representing the total #preK kids each day
	'''
	for i in range(len(totals)):
		if int(olderTotals[i]) + int(preKTotals[i]) != int(totals[i][1]):
			raise Exception("The totals are wrong for", totals[i][0])


def get_groups(lines):
	group1 = True

	olderGroup = []
	preKGroup = []
	olderTotals = []
	preKTotals = []

	i = 0
	while i < len(lines):
		line = lines[i].strip()
		line = line.split("\t")
		if group1:
			if "Total" in line[0]:
				# store the older kid totals for verification, and skip ahead to the next section
				olderTotals = line[1::]
				group1 = False
				i += 4
			else:
				olderGroup.append(line)
				i += 1
		else:
			if "Total" in line[0]:
				# store the younger kid totals for verification, break
				preKTotals = line[1::]
				break
			else:
				preKGroup.append(line)
				i += 1

	return (olderGroup, preKGroup, olderTotals, preKTotals)


def writeToSheets(days, olderTotals, preKTotals, olderGroup, preKGroup, sundayDate):
	'''
	Driver code to write Older and PreK data into their respective workbooks and sheets

	Args:
		days: a tuple array representing the days and whole program
			totals for each day. Format is (day, total)
		olderTotals: int array representing the older group's totals for each day
		preKTotals: int array representing the preK group's totals for each day
		olderGroup: a 2D array of older group kids and days they're planning to show
			up (represented by a 1 or 0). This array is in the format: [name, 1, 0, 1]
		preKGroup: same as `olderGroup`, but for preK kids
		sundayDate: a string of the format mm/dd/yy, representing the week's sunday date
	'''
	# copy files for template
	shutil.copy("CHILD HEADCOUNT TEMP.xlsx", "Older.xlsx")
	shutil.copy("CHILD HEADCOUNT TEMP.xlsx", "PreK.xlsx")

	wbOlder = load_workbook('Older.xlsx')
	wbPreK = load_workbook('PreK.xlsx')


	# get the dates of each of the days for the headcount sheets
	for day in days:
		if day[0] == "Monday":
			day.append(get_next_day(sundayDate, 1))
		elif day[0] == "Tuesday":
			day.append(get_next_day(sundayDate, 2))
		elif day[0] == "Wednesday":
			day.append(get_next_day(sundayDate, 3))
		elif day[0] == "Thursday":
			day.append(get_next_day(sundayDate, 4))
		elif day[0] == "Friday":
			day.append(get_next_day(sundayDate, 5))


	for i in range(len(days)):
		writeToSheet(days, olderTotals, olderGroup, i, wbOlder, "OlderGroup")
		writeToSheet(days, preKTotals, preKGroup, i, wbPreK, "PreKGroup")

	# delete the first sheet in each one
	wbOlder.remove(wbOlder['Sheet1'])
	wbPreK.remove(wbPreK['Sheet1'])

	wbOlder.save('Older.xlsx')
	wbPreK.save('PreK.xlsx')



def get_next_day(prevDay, numDaysLater):
	'''
	Gets a date in mm/dd/yy string format.

	Args:
		prevDay: a string date in mm/dd/yy format, used as reference to get the next date
		numDaysLater: an int representing the number of days from the prevDay you want to get
	'''
	prevDay = prevDay.split("/")

	date = datetime.datetime(int(f"20{prevDay[2]}"), int(prevDay[0]), int(prevDay[1]))
	date += datetime.timedelta(days=numDaysLater)
	return date.strftime("%m/%d/%y")



def writeToSheet(days, totals, group, day, wb, sheetName):
	'''
	Writes one day's worth of kids into 1 sheet, with formatting

	Args:
		days: a tuple array representing the days and whole program
			totals for each day. Format is (day, total)
		totals: int array representing the group's totals for each day
		group: a 2D array of kids and the days they're planning to show
			up (represented by a 1 or 0). This array is in the format: [name, 1, 0, 1]
		day: an int representing the index of the day that the sheet is detailing
		wb: the Workbook object, used for writing to an excel file
		sheetName: title of the sheet, to differentiate older/preK groups
	'''
	target = wb['Sheet1']
	wb.copy_worksheet(target)
	sheet1 = wb.worksheets[-1]
	sheet1.title = f'{sheetName}_{days[day][0]}'

	# add date and week day
	sheet1["H4"] = days[day][0]
	sheet1["O6"] = days[day][2]
	sheet1["J4"] = f"Total: {totals[day]}"

	# add champions logo
	my_png = openpyxl.drawing.image.Image('logo.png')
	my_png.height = 110
	my_png.width = 233
	sheet1.add_image(my_png, 'B1')

	i = 9
	for j, kid in enumerate(group):
		if kid[day+1] == "1":
			# write kid names in
			sheet1[f'G{i}'] = kid[0]
			i += 1

		if i > 34:
			j += 1
			# start new sheet
			wb.copy_worksheet(target)
			sheet2 = wb.worksheets[-1]
			sheet2.title = f'{sheetName}_{days[day][0]}_2'

			# add date and week day
			sheet2["H4"] = days[day][0]
			sheet2["O6"] = days[day][2]

			# add champions logo
			my_png = openpyxl.drawing.image.Image('logo.png')
			my_png.height = 110
			my_png.width = 233
			sheet2.add_image(my_png, 'B1')

			for kiddo in group[j::]:
				if kiddo[day+1] == "1":
					# write kid names in
					sheet2[f'G{i-26}'] = kiddo[0]
					i += 1
			break



def printDailyKids(totals, group):
	'''
	For either the preK or older group, prints kids attending for each day.
	Helper function used for quick code verifications.

	Args:
		totals: an array representing the days and total kids. Needed for the days
		group: an array representing all the kids in either the pre k or older group
	'''
	for i in range(len(totals)):
		day = totals[i][0]
		print(f"*************** {day} ***************")
		for kid in group:
			if kid[i+1] == "1":
				print(kid[0].strip().split(", ")[0], end ="\t")
				print(kid[0].strip().split(", ")[1])



def printAll(totals, group):
	'''
	Prints everything.
	Helper function used for quick code verifications.

	Args:
		totals: an array representing the days and total kids. Needed for the days
		group: an array representing all the kids in either the pre k or older group
	'''
	print("Names", end ="\t\t")
	for i in range(len(totals)):
		print(totals[i][0], end ="\t")
	print()

	for kid in group:
		for val in kid:
			print(val, end ="\t")
		print()



read_data('headcount_data.txt')





